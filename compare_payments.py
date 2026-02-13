import argparse
import os
import re
import warnings
from decimal import Decimal, InvalidOperation
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

import pandas as pd
from openpyxl.styles import PatternFill


ID_BRACKET_RE = re.compile(r"\[(\d+)\]")

# Some source workbooks contain an invalid print-area defined name.
# This warning does not affect data reading, so we hide it in console output.
warnings.filterwarnings(
    "ignore",
    message=r"Print area cannot be set to Defined name: .*",
    category=UserWarning,
)


def parse_amount(value) -> Decimal | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None

    text = str(value).strip().replace(" ", "")
    if not text:
        return None

    # Support both "123,45" and "123.45".
    text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def parse_id_from_file1(name_with_id) -> str | None:
    if name_with_id is None or (isinstance(name_with_id, float) and pd.isna(name_with_id)):
        return None
    match = ID_BRACKET_RE.search(str(name_with_id))
    return match.group(1) if match else None


def parse_id_from_file2(raw_id) -> str | None:
    if raw_id is None or (isinstance(raw_id, float) and pd.isna(raw_id)):
        return None
    text = str(raw_id).strip()
    if not text:
        return None
    return text.split(",")[0].strip()


def prepare_file1(path: Path) -> pd.DataFrame:
    # File 1 layout:
    # - company with ID in column D starting row 9 (in some exports can be C)
    # - amount in column AF
    df = pd.read_excel(
        path,
        sheet_name=0,
        header=None,
        skiprows=8,
        usecols="C,D,AF",
        engine="openpyxl",
    )
    df.columns = ["name_c", "name_d", "amount"]

    # Prefer D by specification, fallback to C when D is empty.
    df["company_raw"] = df["name_d"].where(df["name_d"].notna(), df["name_c"])
    df["company_id"] = df["company_raw"].apply(parse_id_from_file1)
    df["amount_1"] = df["amount"].apply(parse_amount)

    out = df.loc[df["company_id"].notna(), ["company_id", "company_raw", "amount_1"]].copy()
    out["company_id"] = out["company_id"].astype(str).str.strip()
    out["company_name_1"] = out["company_raw"].astype(str).str.strip()

    agg = (
        out.groupby("company_id", as_index=False)
        .agg(
            amount_1=("amount_1", lambda x: sum((v for v in x if v is not None), Decimal("0"))),
            company_name_1=("company_name_1", "first"),
        )
    )
    return agg


def prepare_file2(path: Path) -> pd.DataFrame:
    # File 2 layout:
    # - ID in column A starting row 5, format "*,000" (use part before comma)
    # - company name in column C by specification (in some exports can be B)
    # - amount in column F
    df = pd.read_excel(
        path,
        sheet_name=0,
        header=None,
        skiprows=4,
        usecols="A,B,C,F",
        engine="openpyxl",
    )
    df.columns = ["raw_id", "name_b", "name_c", "amount"]

    df["company_id"] = df["raw_id"].apply(parse_id_from_file2)
    df["company_raw"] = df["name_c"].where(df["name_c"].notna(), df["name_b"])
    df["amount_2"] = df["amount"].apply(parse_amount)

    out = df.loc[df["company_id"].notna(), ["company_id", "company_raw", "amount_2"]].copy()
    out["company_id"] = out["company_id"].astype(str).str.strip()
    out["company_name_2"] = out["company_raw"].astype(str).str.strip()

    agg = (
        out.groupby("company_id", as_index=False)
        .agg(
            amount_2=("amount_2", lambda x: sum((v for v in x if v is not None), Decimal("0"))),
            company_name_2=("company_name_2", "first"),
        )
    )
    return agg


def compare(file1: Path, file2: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    left = prepare_file1(file1)
    right = prepare_file2(file2)
    file1_name = file1.name
    file2_name = file2.name

    merged = left.merge(right, on="company_id", how="outer")
    merged["amount_1"] = merged["amount_1"].fillna(Decimal("0"))
    merged["amount_2"] = merged["amount_2"].fillna(Decimal("0"))
    merged["delta"] = merged["amount_1"] - merged["amount_2"]

    def get_status(row) -> str:
        has_1 = pd.notna(row.get("company_name_1"))
        has_2 = pd.notna(row.get("company_name_2"))
        if has_1 and has_2:
            return "MATCH" if row["delta"] == 0 else "MISMATCH"
        if has_1:
            return f"only_file_{file1_name}"
        return f"only_file_{file2_name}"

    merged["status"] = merged.apply(get_status, axis=1)
    # Final report order: ascending by column A (company_id).
    merged = merged.sort_values(by=["company_id"]).reset_index(drop=True)

    detail = merged[
        [
            "company_id",
            "company_name_1",
            "company_name_2",
            "amount_1",
            "amount_2",
            "delta",
            "status",
        ]
    ].copy()

    summary = (
        detail.groupby("status", as_index=False)
        .agg(
            rows=("company_id", "count"),
            total_amount_1=("amount_1", lambda x: sum(x, Decimal("0"))),
            total_amount_2=("amount_2", lambda x: sum(x, Decimal("0"))),
            total_delta=("delta", lambda x: sum(x, Decimal("0"))),
        )
        .sort_values(by="status")
    )

    detail = detail.rename(
        columns={
            "company_name_1": f"company_name_{file1.name}",
            "company_name_2": f"company_name_{file2.name}",
            "amount_1": f"amount__{file1.name}",
            "amount_2": f"amount__{file2.name}",
        }
    )

    return detail, summary


def to_numeric_columns(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    out = df.copy()
    for col in cols:
        out[col] = out[col].astype(str).astype(float)
    return out


def write_report(detail: pd.DataFrame, summary: pd.DataFrame, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    detail_amount_cols = [col for col in detail.columns if col.startswith("amount__")]
    detail_out = to_numeric_columns(detail, detail_amount_cols + ["delta"])
    summary_out = to_numeric_columns(
        summary, ["total_amount_1", "total_amount_2", "total_delta"]
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        detail_out.to_excel(writer, sheet_name="comparison", index=False)
        summary_out.to_excel(writer, sheet_name="summary", index=False)

        match_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        other_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")

        for sheet_name in ("comparison", "summary"):
            ws = writer.book[sheet_name]
            headers = [cell.value for cell in ws[1]]
            if "status" not in headers:
                continue
            status_col_idx = headers.index("status") + 1

            for row_idx in range(2, ws.max_row + 1):
                status_value = ws.cell(row=row_idx, column=status_col_idx).value
                fill = match_fill if status_value == "MATCH" else other_fill
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        # Comparison columns D/E/F contain monetary values.
        ws_cmp = writer.book["comparison"]
        number_format = "#,##0.00"
        for row_idx in range(2, ws_cmp.max_row + 1):
            for col_idx in (4, 5, 6):
                ws_cmp.cell(row=row_idx, column=col_idx).number_format = number_format

        # Fit column widths to visible content for easier reading.
        for sheet_name in ("comparison", "summary"):
            ws = writer.book[sheet_name]
            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    value = "" if cell.value is None else str(cell.value)
                    if len(value) > max_len:
                        max_len = len(value)
                ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

        # Keep headers on row 1 and enable Excel filters for convenient viewing.
        for sheet_name in ("comparison", "summary"):
            ws = writer.book[sheet_name]
            ws.auto_filter.ref = ws.dimensions


def open_report(path: Path) -> None:
    # Windows: open with the default associated application (typically Excel).
    os.startfile(str(path))


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Compare client payment amounts between two XLSX exports by personal ID."
    )
    parser.add_argument(
        "--file1",
        default="work/OC.xlsx",
        help="Path to the first XLSX file (default: work/OC.xlsx)",
    )
    parser.add_argument(
        "--file2",
        default="work/1C.xlsx",
        help="Path to the second XLSX file (default: work/1C.xlsx)",
    )
    parser.add_argument(
        "--out",
        default="reports/payments_comparison.xlsx",
        help="Output XLSX report path (default: reports/payments_comparison.xlsx)",
    )
    parser.add_argument(
        "--gui",
        action="store_true",
        help="Launch a simple GUI to pick file1, file2 and output report path.",
    )
    return parser


def run_compare(file1: Path, file2: Path, out: Path) -> dict[str, int]:
    detail, summary = compare(file1, file2)
    write_report(detail, summary, out)
    open_report(out)

    mismatch_count = int((detail["status"] == "MISMATCH").sum())
    only1_count = int((detail["status"] == f"only_file_{file1.name}").sum())
    only2_count = int((detail["status"] == f"only_file_{file2.name}").sum())
    return {
        "rows": len(detail),
        "mismatches": mismatch_count,
        "only1": only1_count,
        "only2": only2_count,
    }


def launch_gui(default_file1: str, default_file2: str, default_out: str) -> None:
    root = tk.Tk()
    root.title("Сверка платежей")
    root.geometry("880x220")
    root.resizable(False, False)

    file1_var = tk.StringVar(value=default_file1)
    file2_var = tk.StringVar(value=default_file2)
    out_var = tk.StringVar(value=default_out)

    def pick_file1() -> None:
        path = filedialog.askopenfilename(
            title="Выберите первый файл",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            file1_var.set(path)

    def pick_file2() -> None:
        path = filedialog.askopenfilename(
            title="Выберите второй файл",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            file2_var.set(path)

    def pick_out() -> None:
        path = filedialog.asksaveasfilename(
            title="Куда сохранить отчет",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=Path(out_var.get()).name if out_var.get() else "payments_comparison.xlsx",
        )
        if path:
            out_var.set(path)

    def generate() -> None:
        file1 = Path(file1_var.get().strip())
        file2 = Path(file2_var.get().strip())
        out = Path(out_var.get().strip())

        if not file1_var.get().strip() or not file2_var.get().strip() or not out_var.get().strip():
            messagebox.showerror("Ошибка", "Заполните все три пути: файл 1, файл 2 и выходной отчет.")
            return
        if not file1.is_file():
            messagebox.showerror("Ошибка", f"Первый файл не найден:\n{file1}")
            return
        if not file2.is_file():
            messagebox.showerror("Ошибка", f"Второй файл не найден:\n{file2}")
            return

        try:
            stats = run_compare(file1, file2, out)
        except Exception as exc:
            messagebox.showerror("Ошибка", str(exc))
            return

        messagebox.showinfo(
            "Готово",
            (
                f"Отчет сохранен:\n{out}\n\n"
                f"Rows: {stats['rows']}\n"
                f"Mismatches: {stats['mismatches']}\n"
                f"only_file_{file1.name}: {stats['only1']}\n"
                f"only_file_{file2.name}: {stats['only2']}"
            ),
        )

    root.columnconfigure(1, weight=1)

    tk.Label(root, text="Укажите файл OC:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
    tk.Entry(root, textvariable=file1_var).grid(row=0, column=1, padx=10, pady=10, sticky="ew")
    tk.Button(root, text="Выбрать...", command=pick_file1, width=14).grid(row=0, column=2, padx=10, pady=10)

    tk.Label(root, text="Укажите файл 1С:").grid(row=1, column=0, padx=10, pady=10, sticky="w")
    tk.Entry(root, textvariable=file2_var).grid(row=1, column=1, padx=10, pady=10, sticky="ew")
    tk.Button(root, text="Выбрать...", command=pick_file2, width=14).grid(row=1, column=2, padx=10, pady=10)

    tk.Label(root, text="Файл отчета:").grid(row=2, column=0, padx=10, pady=10, sticky="w")
    tk.Entry(root, textvariable=out_var).grid(row=2, column=1, padx=10, pady=10, sticky="ew")
    tk.Button(root, text="Сохранить как...", command=pick_out, width=14).grid(row=2, column=2, padx=10, pady=10)

    tk.Button(root, text="Сверить банк", command=generate, width=20).grid(
        row=3, column=1, padx=10, pady=20, sticky="e"
    )

    root.mainloop()


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    if args.gui:
        launch_gui(args.file1, args.file2, args.out)
        return

    file1 = Path(args.file1)
    file2 = Path(args.file2)
    out = Path(args.out)

    stats = run_compare(file1, file2, out)
    print(f"Saved: {out}")
    print(
        f"Rows: {stats['rows']} | mismatches: {stats['mismatches']} | "
        f"only_file_{file1.name}: {stats['only1']} | only_file_{file2.name}: {stats['only2']}"
    )


if __name__ == "__main__":
    main()
